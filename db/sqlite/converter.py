from openpyxl import load_workbook
import yaml
import fs
from regx import Regx
import sqlite3
from data.schema import DataManager

version = '1.0'
DB_SCHEMA_HEADER = r"""
Root: 
    class: dict
    keys:
    - {name: 'Type', rule: 'Type', required: True}
    - {name: 'Version', rule: 'Version', required: True}
    - {name: 'Comment', rule: 'Comment', required: False}
    root: True
Type: 
    class: str
    equals: SQLiteDatabaseObject
Version:
    class: str
Comment:
    class: str
"""
DB_SCHEMA_TABLE = r"""
TableName:
    class: dict
    keys:
    - {name: '(\w+)', regx: True, rule: 'TableObject', required: False, min-count: 1}
    root: True
TableObject:
    class: dict
    - {name: 'ColumnInfo', rule: 'ColumnInfo', required: True}
    - {name: 'RowData', rule: 'DataRows', required: True}
ColumnInfo
    class: dict
    keys: 
    - {name: 'Order', rule: 'ColumnOrder', required: True}
    - {name: 'Definition', rule: 'ColumnDefinition', required: True}
ColumnOrder:
    class: list
    render: {yaml: {collapse: True}}   
ColumnDefinition:
    class: dict
    keys: 
    - {name: '(\w+)', regx: True, rule: 'ColumnDefinitionObject', required: False, min-count: 1}
ColumnDefinitionObject:
    class: dict
    keys:
    - {name: 'Index', rule: 'Index', required: True}
    - {name: 'Type', rule: 'ColumnType', required: True}
    - {name: 'Attributes', rule: 'Attributes', required: True, Default: {}}
    render: {yaml: {collapse: True}}   
Index:
    class: int
ColumnType:
    class: str
    in: ['text', 'integer', 'real']
Attributes:
    class: dict
    keys:
    - {name: 'PrimaryKey', rule: 'Bool', required: False}
    - {name: 'ForeignKey', rule: 'ForeignKey', required: False}
    - {name: 'NotNull', rule: 'Bool', required: False}
Bool:
    class: bool
ForeignKey:
    class: str
"""

other = r"""
TableObject:
  class: dict
  keys:
  - {name: 'ColumnInfo', rule: 'ColumnInfo', required: True}
  - {name: 'DataRows', rule: 'DataRows', required: True}
  - {name: 'Comment', rule: 'Comment', required: False}
ColumnInfo:
  class: dict
  - {name: 'ColumnOrder', rule: 'ColumnOrder', required: True}
  - {name: 'ColumnDefinition', rule: 'ColumnDefinition', required: True}
DataRows:
  class: list
  rule: DataRow
DataRow: {}
"""

def convert_bool_to_int(val):
    t = type(val)
    if t == str:
        new_val = val.lower()
        if new_val == 'true': return 1
        if new_val == 'false': return 0
        regx = Regx()
        if regx.m(new_val, r'^t'): return 1
        else: return 0
    try:
        new_val = bool(val)
    except:
        raise Exception('Could not convert "{}" to type bool'.format(val))
    return(new_val)

class SQLiteConvert():

    def __init__(self):
        self.header = \
        {
            'Type': 'SQLiteDatabaseObject',
            'Version': version,
        }
        self.data = {}

    def read_xlsx(self, xlsx_file, verbose=1):
        if verbose > 0: print('Reading "{}" ...'.format(xlsx_file))
        re = Regx()
        wb = load_workbook(filename = xlsx_file, data_only = True)
        data = {}
        for sheet in wb.worksheets:
            sheet_name = sheet.title
            if verbose > 0: print('- Parsing sheet {}'.format(sheet_name))
            data[sheet_name] = {'ColumnNames': [], 'ColumnInfo': {}, 'RowData': []}
            headers = {}
            val_type = {}
            # Get headers
            for c in range(sheet.min_column, sheet.max_column+1):
                r = sheet.min_row
                header = sheet.cell(r, c).internal_value
                data[sheet_name]['ColumnNames'].append(header)
                val_type[header] = {}
                headers[c] = header
            # Read data from sheet.  Also, record best guess as to data type.
            for r in range(sheet.min_row+1, sheet.max_row+1):
                row = {}
                for c in range(sheet.min_column, sheet.max_column+1):
                    val = str(sheet.cell(r, c).value)
                    header = headers[c]
                    row[header] = val
                    if val.lower() == 'true' or val.lower() == 'false':
                        val_type[header]['bool'] = True
                    elif re.m(val, r'^[-+]?\d+$', ''): 
                        val_type[header]['int'] = True
                    elif re.m(val, r'^[-+]?\d*\.\d+(?:[eE][-+]?\d+){0,1}$', ''):
                        val_type[header]['float'] = True
                    else:
                        val_type[header]['str'] = True
                data[sheet_name]['RowData'].append(row)
            # Get cast functions.
            recast_func = {}
            for c in headers:
                header = headers[c]
                data[sheet_name]['ColumnInfo'][header] = {}
                if 'str' in val_type[header]:
                    data[sheet_name]['ColumnInfo'][header] = {'Index': c, 'Type': 'text', 'Definition': {}}
                    recast_func[header] = str
                elif 'float' in val_type[header]:
                    data[sheet_name]['ColumnInfo'][header] = {'Index': c, 'Type': 'real', 'Definition': {}}
                    recast_func[header] = float
                elif 'int' in val_type[header]:
                    data[sheet_name]['ColumnInfo'][header] = {'Index': c, 'Type': 'integer', 'Definition': {}}
                    recast_func[header] = int
                elif 'bool' in val_type[header] and len(val_type[header]) == 1:
                    data[sheet_name]['ColumnInfo'][header] = {'Index': c, 'Type': 'integer', 'Definition': {}}
                    recast_func[header] = convert_bool_to_int
                else:
                    data[sheet_name]['ColumnInfo'][header] = {'Index': c, 'Type': 'text', 'Definition': {}}
                    recast_func[header] = str
            # Re-cast data values using cast functions.
            for row in data[sheet_name]['RowData']:
                for c in headers:
                    header = headers[c]
                    row[header] = recast_func[header](row[header])
        self.data = data
        return(data)

    def validate(self, data=None):
        if data is None: data = self.data
        schema = yaml.load(DB_SCHEMA_HEADER, Loader=yaml.FullLoader)
        dm = DataManager(schema)
        dm.validate(self.header)
        for table in data: pass
        ### fix - left off here
        return(True)

    def write_yaml(self, yaml_file, data=None, verbose=1):
        self.validate()
        if data is None: data = self.data
        schema = yaml.load(DB_SCHEMA_HEADER)
        print(schema)
        return
        if verbose > 0: print('Writing "{}" ...'.format(yaml_file))
        yaml_text = yaml.dump(data)
        if verbose > 0: print('- ' + fs.write_file_if_changed(yaml_file, yaml_text))
        else: fs.write_file('- ' + yaml_file, yaml_text)

    def read_yaml(self, yaml_file, verbose=1):
        if verbose > 0: print('Reading "{}" ...'.format(yaml_file))
        yaml_text = fs.read_file(yaml_file)
        data = yaml.load(yaml_text)
        return(data)

    def write_sqlite(self, db_file, data, verbose=1):
        if fs.file_exists(db_file): fs.delete_file(db_file)
        if verbose > 0: print('Creating "{}" ...'.format(db_file))
        db = sqlite3.connect(db_file)
        def bool_to_int(val):
            val = str(val).lower()
            if val == 'true': return 1
            if val == 'false': return 0
            return int(bool(val))
        xlsx_type = \
        {
            'int': {'name': 'integer', 'func': int},
            'str': {'name': 'text', 'func': str},
            'float': {'name': 'real', 'func': float},
            'bool': {'name': 'integer', 'func': bool_to_int},
        }
        for table in data:
            cols = []
            func = {}
            for header in data[table]['ColumnNames']:
                db_type = xlsx_type[data[table]['ColumnInfo'][header]]['name']
                func[header] =  xlsx_type[data[table]['ColumnInfo'][header]]['func']
                cols.append('"{}" {}'.format(header, db_type))
            sql_create_table = 'create table {} ({})'.format(table, ', '.join(cols))
            curs = db.cursor()
            curs.execute(sql_create_table)
            for row in data[table]['RowData']:
                cols = []
                vals = []
                bindings = []
                for header in data[table]['ColumnNames']:
                    cols.append('"{}"'.format(header))
                    val = func[header](row[header])
                    vals.append('?')
                    bindings.append(val)
                sql_insert = 'insert into {} ({}) values ({})'.format(table, ', '.join(cols), ', '.join(vals))
                curs.execute(sql_insert, bindings)
            db.commit()
        db.close()

